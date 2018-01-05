# -*- coding: utf-8 -*-

'''
Created on 7 apr. 2017

@author: perhk
'''



import json, csv

from django.http import HttpResponse, HttpResponseNotFound, HttpResponseRedirect
from django.shortcuts import render

import os
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# gcash = {}
# def fill_gcash():
#     reader = csv.DictReader(open(BASE_DIR+"/geocodecash.csv"))
#     for row in reader:
#         gcash[row['Adress']] = [row['Lat'],row['Lon']]
# 
# def get_gcode(adress):
#     if adress in gcash:
#         return gcash[adress][0],gcash[adress][1]
#     else:
#         return 0,0
# 
# mcash = []
# def fill_mcash():
#     avdelningar = {"Sagodjuren":"Spårare","Husdjuren":"Spårare","Gosedjuren":"Spårare","Nya spårare att fördela":"Spårare","Fabeldjuren":"Upptäckare","Skogsdjuren":"Upptäckare","Urdjuren":"Äventyrare","Rovdjuren":"Äventyrare","Slow Fox":"Utmanare","Rover":"Rover","Övriga ledare":"Ledare"}
#     if len(gcash) == 0:
#         fill_gcash()
#     reader = csv.DictReader(open(BASE_DIR+"/medlemslista.csv"))
#     for row in reader:
#         avd = row['Avdelning']
#         if avd in avdelningar: 
#             namn = row['Förnamn']+" "+row['Efternamn']
#             if row['Födselsdatum'][0:4] < "1992":
#                 grupp = "Ledare"
#             else:
#                 grupp = avdelningar[avd]
#             adr = row['Adress']+", "+row['Postnr']+" "+row['Postort']+", "+row['Land']
#             lat,lon = get_gcode(adr)
#             mcash.append({"namn":namn,"grupp":grupp,"avdelning":avd,"position":(lat,lon)})

# import dropbox
from openpyxl import load_workbook
import requests
from io import BytesIO

fibcash = []
def fill_fibcash():
    fiberfile = "https://www.dropbox.com/s/wuipbqguzyc5ouv/Fiberbest%C3%A4llning.xlsx?dl=1"
    r = requests.get(fiberfile)
    wb = load_workbook(BytesIO(r.content), read_only=True)
#     wb = load_workbook(filename = BASE_DIR+"/Fiberbeställning.xlsx", read_only=True)
    ws = wb['fiber']
    head = True
    for row in ws.rows:
        if not head:
            v1 = row[0].value
            v2 = row[1].value
            v3 = row[2].value
            v4 = row[3].value
            if v1 == None or v2 == None:
                break
            fibcash.append({"fastighet":v1,"adress":v2,"status":v4,"position":v3.split(",")})
        else:
            head = False
            

def bing(request):
    return render(request, 'bing.html')

# def index(request):
#     if len(mcash) == 0:
#         fill_mcash()
#     return render(request, 'index.html')

def karta(request):
#     if len(fibcash) == 0:
    fibcash = []
    fill_fibcash()
    return render(request, 'karta.html')


def places(request):
    lat1=request.GET['lat1'][:8]
    lon1=request.GET['lon1'][:8]
    lat2=request.GET['lat2'][:8]
    lon2=request.GET['lon2'][:8]
    ret = []
    for f in fibcash:
        lat,lon = f['position']
#         if lat < lat1 and lat >lat2 and lon < lon1 and lon > lon2:
        ret.append({"fastighet":f['fastighet'],"adress":f['adress'],"status":f['status'],"position":[lat,lon]})
    return HttpResponse(json.dumps(ret), content_type='application/json')




